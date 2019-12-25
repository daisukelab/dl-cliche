from dlcliche.utils import *
import openpyxl as opx
try:
    from openpyxl.worksheet import Worksheet
except:
    from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from copy import copy
import re


def opx_copy_cell(source_cell, target_cell, style_copy=False):
    """openpyxl helper: Copy cell from source to target.
    Cells can be on different sheets.

    Arguments:
        style_copy: Flag to copy style or not.
    """
    target_cell._value = source_cell._value
    target_cell.data_type = source_cell.data_type
    if style_copy and source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def opx_copy_cells(src_sh, src_pos, dest_sh, dest_pos, row_step=1, col_step=0, stop_if_empty=True, style_copy=False):
    """Copy cells between sheets sequentially.

    Arguments:
        src_sh: Source sheet
        src_pos: Source position (column, row)
        dest_sh: Destination sheet
        dest_pos: Destination position
        row_step: Row step to move copy position; 2 will copy like this: B2, B4, ...
        col_step: Column step; 2 will work like this: B2, D2, ...
        stop_if_empty: Stop copying if source cell is empty, i.e. str(cell) == ''
        style_copy: Copy cell style or not

    Returns:
        Number of cells copied.
    """
    src, dst = src_sh, dest_sh
    src_col, src_row = src_pos
    dst_col, dst_row = dest_pos
    count = 0
    while True:
        opx_copy_cell(src.cell(column=src_col+1, row=src_row+1),
                      dst.cell(column=dst_col+1, row=dst_row+1),
                      style_copy=style_copy)
        src_col += col_step
        dst_col += col_step
        src_row += row_step
        dst_row += row_step
        count += 1
        if src.max_column <= src_col: break
        if src.max_row    <= src_row: break
        if stop_if_empty and str(src.cell(column=src_col+1, row=src_row+1)) == '': break
    return count


class _ForcedWorksheetCopy(object):
    """
    Custom made copier.
    Thanks to https://bitbucket.org/openpyxl/openpyxl/src/d090796b358b0238cce5e9c71fa2dd1ccc37ccbc/openpyxl/worksheet/copier.py?at=default&fileviewer=file-view-default
    """

    def __init__(self, source_worksheet, target_worksheet):
        self.source = source_worksheet
        self.target = target_worksheet
        self._verify_resources()

    def _verify_resources(self):

        if (not isinstance(self.source, Worksheet)
            and not isinstance(self.target, Worksheet)):
            raise TypeError("Can only copy worksheets")

        if self.source is self.target:
            raise ValueError("Cannot copy a worksheet to itself")

        # Enabled this --> if self.source.parent != self.target.parent:
        #    raise ValueError('Cannot copy between worksheets from different workbooks')

    def copy_worksheet(self):
        self._copy_cells()
        self._copy_dimensions()

        self.target.sheet_format = copy(self.source.sheet_format)
        self.target.sheet_properties = copy(self.source.sheet_properties)
        self.target.merged_cells = copy(self.source.merged_cells)

    def _copy_cells(self):
        for (row, col), source_cell  in self.source._cells.items():
            target_cell = self.target.cell(column=col, row=row)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                opx_copy_cell_style(source_cell, target_cell,
                                    between_wb=(self.source.parent != self.target.parent))

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    def _copy_dimensions(self):
        for attr in ('row_dimensions', 'column_dimensions'):
            src = getattr(self.source, attr)
            target = getattr(self.target, attr)
            for key, dim in src.items():
                target[key] = copy(dim)
                target[key].worksheet = self.target


def opx_copy_worksheet(source_worksheet, target_worksheet):
    """Copy worksheet even in between workbooks."""
    copier = _ForcedWorksheetCopy(source_worksheet=source_worksheet, target_worksheet=target_worksheet)
    copier.copy_worksheet()


def opx_color_cell(cell, rgb='00FF0000', pattern_type='solid', negative_rgb=None, keywords=None):
    """openpyxl helper: Set cell color and its pattern.
    If keywords are listed, apply only when any keyword is in value string,
    and if negative_rgb is also set, this color will be set if no keyword matches.
    Thanks to https://stackoverflow.com/questions/30484220/python-fill-cells-with-colors-using-openpyxl
    
    Arguments:
        cell: The cell to set color.
        rgb: Cell color.
        pattern_type: Cell filling pattern.
        negative_rgb: Valid only keywords are listed, set this color when no keyword maches.
        keywords: List of keyword texts. Color will be set if any of keyword matches.

    Returns:
        Color was set (True) or not (False) when keywords are listed.
        Always True if no keywords being set.
    """
    if keywords is not None:
        text = str(cell._value)
        apply_color = False
        for k in keywords:
            if re.search(k, text) is not None:
                apply_color = True
                break
        if not apply_color and negative_rgb is not None:
            apply_color = True
            rgb = negative_rgb
    else:
        apply_color = True

    if apply_color:
        this_color = opx.styles.colors.Color(rgb=rgb)
        this_fill = opx.styles.fills.PatternFill(patternType=pattern_type, fgColor=this_color)
        cell.fill = this_fill
    return apply_color


def opx_color_rows(worksheet, fn_cell_rgb, context=None, pattern_type='solid'):
    """openpyxl helper: Set cell color for each rows.
    
    Arguments:
        worksheet: The worksheet to run through.
        fn_cell_rgb: Function to determine which column in the row to have what cell color.
            `fn_cell_rgb(worksheet, row:int, num_of_columns, context=context): list(list())`
            supposed to return 2D array. TODO explain more.
        pattern_type: Cell filling pattern.
    """
    n_col = worksheet.max_column
    n_row = worksheet.max_row
    for r in range(n_row):
        col_rgbs = fn_cell_rgb(worksheet, r, n_col, context=context)
        for c, rgb in col_rgbs:
            this_color = opx.styles.colors.Color(rgb=rgb)
            this_fill = opx.styles.fills.PatternFill(patternType=pattern_type, fgColor=this_color)
            worksheet.cell(column=c+1, row=r+1).fill = this_fill

def opx_copy_cell_style(source_cell, target_cell, between_wb=False):
    """openpyxl helper: Copy cell style only from source to target.
    Thanks to https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
    """
    if between_wb:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
    else:
        # Much faster and reliable
        target_cell._style = copy(source_cell._style)


def opx_copy_row(sh_src, row_src, sh_dest, row_dest, n_col=None, style_copy=False, debug=False):
    """openpyxl helper: Copy row.

    Arguments:
        sh_src: Source sheet.
        row_src: Source row number, range is as usual: `[0, max-1]`.
        sh_dest: Destination sheet.
        row_dest: Destination row number, same range as source.
        n_col: Number of columns to copy, n_col=5 will copy five sequential columns.
        style_copy: Flag to copy style or not.
    """
    if n_col is None: n_col = sh_src.max_column
    if debug:
        print('<', 1, row_src+1, sh_src.cell(column=1, row=row_src+1).value)
        print('>', 1, row_dest+1, sh_dest.cell(column=1, row=row_dest+1).value)
    for c in range(n_col):
        opx_copy_cell(sh_src.cell(column=c+1, row=row_src+1), sh_dest.cell(column=c+1, row=row_dest+1),
                     style_copy=style_copy)


def opx_duplicate_style(sh, row_src, row_dest, n_row=None, debug=False):
    """openpyxl helper: Copy style among rows.

    Arguments:
        sh: Sheet.
        row_src: Source row number, range is as usual: `[0, max-1]`.
        row_dest: Destination row number, same range as source.
        n_row: Number of rows to copy, n_col=5 will copy five sequential rows.
        style_copy: Flag to copy style or not.
    """
    n_col = sh.max_column
    if n_row is None:
        n_row = sh.max_row - row_dest
    if debug:
        print('duplicate', row_src, 'style to', row_dest, '-', row_dest+n_row - 1)
    for r in range(row_dest, row_dest+n_row):
        for c in range(n_col):
            opx_copy_cell_style(sh.cell(column=c+1, row=row_src+1), sh.cell(column=c+1, row=r+1))


def opx_reorder_sheets(wb, name_list):
    """Reorder worksheets of workbook object.
    Thanks to https://stackoverflow.com/questions/51082458/move-a-worksheet-in-a-workbook-using-openpyxl-or-xl-or-xlsxwriter
    """
    names = wb.sheetnames.copy()
    assert set(names) == set(name_list)
    new_order = [names.index(s) for s in name_list]
    wb._sheets = [wb._sheets[o] for o in new_order]


def opx_set_active_by_name(wb, sheetname):
    """Set worksheet of workbook active by name."""
    wb.active = wb.sheetnames.index(sheetname)


def opx_drop_all_except_active_sheet(wb, keep=[]):
    """Drop all worksheets except active one.
    
    Arguments:
        keep: List of exception sheetnames. All sheets on this list will not be dropped.
    """
    sheet_to_keep = [wb.active.title] + keep
    for sh in wb.sheetnames:
        if sh not in sheet_to_keep:
            del wb[sh]
    wb.active = 0


opx_ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\000-\010]|[\013-\014]|[\016-\037]|[\x7f-\x9f]|[\uffff]')
def opx_remove_illegal_char(data):
    """Remove ILLEGAL CHARACTERS.
    Thanks to https://qiita.com/analytics-hiro/items/e6581019a11de798002b
    but fixed problem to remove new lines... Removed: `\x00-\x1f`.
    """
    if isinstance(data, str):
        return opx_ILLEGAL_CHARACTERS_RE.sub("", data)
    else:
        return data


def opx_df_to_ws(workbook, sheetname, df, start_row=1, start_col=1, index=True, header=True, index_filter=None):
    """Write all data in a DataFrame to Excel worksheet.
    
    Arguments:
        workbook: Target workbook object.
        sheetname: Target worksheet name, will be created if not there.
        df: DataFrame object to read from.
        start_row: Target worksheet row to start writing.
        start_col: Target worksheet column to start writing.
        index: Put index data also if this is True.
        header: Put columns as title row on top of data if this is True.
        index_filter: Function to filter index value like `lambda x: x[:5]` to limit up to 5 letters.

    Returns:
        n_row: Number of rows written, regardless of starting position.
        n_col: Number of columns in the same fashion.
    """
    # Get existing or new worksheet
    if sheetname in workbook.sheetnames:
        ws = workbook[sheetname]
    else:
        ws = workbook.create_sheet(title=sheetname)

    # Put header manually, avoiding two lines...
    n_row, n_col = 0, 0
    if header:
        ws.cell(row=start_row, column=start_col, value=df.index.name)
        header_start_col = start_col+1 if index else start_col
        for ci, value in enumerate(df.columns, header_start_col):
            ws.cell(row=start_row, column=ci, value=opx_remove_illegal_char(value))
        start_row += 1
        n_row += 1
    # Write all data into the sheet
    for ri, row in enumerate(dataframe_to_rows(df, index=index, header=False), start_row):
        if header and index: # index header appears twice, This happens only when header && index.
            if ri == start_row:
                continue
            else:
                ri -= 1
        for ci, value in enumerate(row, start_col):
            if index_filter is not None and ci == start_col:
                value = index_filter(value)
            ws.cell(row=ri, column=ci, value=opx_remove_illegal_char(value))
            if n_col < ci+1: n_col = ci+1
        if n_row < ri: n_row = ri
    return n_row, n_col


def opx_bar_chart(dest_ws, RC, src_ws, TL, BR, figtitle='', xtitle='', ytitle='', figsize=None):
    """Draw bar chart of source data placed in a rectangle area.
    Leftmost column is index (= category).
    
    Arguments:
        dest_ws: Worksheet to append bar chart.
        RC: Row and column position to append bar chart.
        src_ws: Worksheet where chart data are.
        TL: Top left cell `(row, column)` of data area.
        BR: Bottom right cell of data area.
        figtitle: Chart title.
        xtitle: X title.
        ytitle: Y title.
        figsize: Size `(width, height)` of chart.
    """
    chart = opx.chart.BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.title = figtitle
    chart.y_axis.title = ytitle
    chart.x_axis.title = xtitle

    cats = opx.chart.Reference(src_ws, min_col=TL[1], min_row=TL[0], max_col=TL[1], max_row=BR[0])
    data = opx.chart.Reference(src_ws, min_col=TL[1]+1, min_row=TL[0], max_col=BR[1], max_row=BR[0])
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    if figsize is not None:
        chart.width = figsize[0]
        chart.height = figsize[1]
    
    chart.legend = None

    anchor = get_column_letter(RC[1]) + str(RC[0])
    dest_ws.add_chart(chart, anchor)


def _get_type_formula():
    """Wrap version difference.""" 
    try:
        return opx.cell.cell.TYPE_FORMULA  # 2.6.2 ?
    except:
        return opx.cell.Cell.TYPE_FORMULA  # before?


def opx_auto_adjust_column_width(worksheet, max_width=200, default_width=8, scaling=1.1, dont_narrower=False):
    """Auto adjust all column width in a worksheet.
    
    Arguments:
        worksheet: Worksheet object to adjust in place.
        max_width: Max width to limit, or setting None will not limit.
        default_width: Default width, used only when dont_narrower is True.
        scaling: Scaling factor to calculate width; `unicode_visible_width(unistr) * scaling`.
        dont_narrower: Set True if you don't want make columns get narrower.
    """
    column_widths = []
    cell_TYPE_FORMULA = _get_type_formula()
    for row in worksheet.iter_rows():
        for i, cell in enumerate(row):
            if len(column_widths) <= i:
                column_widths.append(default_width)
            if cell_TYPE_FORMULA != cell.data_type:
                column_widths[i] = max(column_widths[i], unicode_visible_width(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        # https://groups.google.com/forum/#!topic/openpyxl-users/ROYUQyH50ro
        cur_width = worksheet.column_dimensions[get_column_letter(i + 1)].width
        if cur_width is None: cur_width = default_width

        if max_width is not None:
            column_width = max_width if max_width < column_width else column_width

        if dont_narrower:
            column_width = cur_width if column_width < cur_width else column_width

        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width * scaling


def opx_set_column_width(worksheet, col, width):
    """Set column absolute width.
    Args:
        col: column number in [1..max_column]
    """
    worksheet.column_dimensions[get_column_letter(col)].width = width


MAX_EXCEL_COL_WIDTH = 20
def df_to_excel_workbook(df, wb=None, template=None, max_col_width=None, ws_name='untitled',
                         index=True, header=True, index_filter=None,
                         view_left_col=None, copy_style=True):
    """Write df to Excel workbook object.
    Refer to df_to_xlsx for the detail.
    """
    # Remember using active sheet or not
    use_active_sheet = wb is None
    # Use/open new/open template workbook
    wb = wb or opx.Workbook()
    if template: wb = opx.load_workbook(template)
    # Set sheetname if new or template
    if use_active_sheet:
        wb.active.title = ws_name
    # Set contents
    opx_df_to_ws(wb, ws_name, df=df, start_row=1, start_col=1, index=index, header=header, index_filter=index_filter)
    if copy_style:
        opx_duplicate_style(wb[ws_name], row_src=2, row_dest=3,
                        n_row=len(df)-3+1, debug=False)
    else:
        opx_auto_adjust_column_width(wb[ws_name], max_width=max_col_width, dont_narrower=False)
    
    # Move active cell to last row
    for i in range(len(wb[ws_name].sheet_view.selection)):
        if wb[ws_name].sheet_view.pane is not None:
            # Set to active pane only
            if wb[ws_name].sheet_view.pane.activePane != wb[ws_name].sheet_view.selection[i].pane:
                continue
        # Set active cell
        wb[ws_name].sheet_view.selection[i].activeCell = f'A{len(df)+1}'
        wb[ws_name].sheet_view.selection[i].sqref = wb[ws_name].sheet_view.selection[i].activeCell
    if view_left_col:
        wb[ws_name].sheet_view.pane.topLeftCell = f'{view_left_col}{int(np.max([len(df)-3, 2]))}'

    return wb


def df_to_xlsx(df, folder, stem_name, template=None, max_col_width=None,
               ws_name=None, index=True, header=True, index_filter=None, view_left_col=None):
    """Write df to Excel .xlsx file.
    Column width will be adjusted to fit the contents.
    Active cell will be set to the top column of last row.

    Arguments:
        df: DataFrame to write.
        folder: Destination folder to place writing file.
        stem_name: File stem; 'abc' will set file name as 'abc.xlsx'.
        template: Template workbook filename.
        max_col_width: (Valid when template is None) Maximum column width
            to prevent column gets too wide. None will not limit width.
        ws_name: Work sheet name. None will set stem_name as work sheet name.
        index_filter: Function to filter index.
        view_left_col: View's leftmost column letter. ex) 'M', 'AL'
    Returns:
        Written path name.
    """
    pathname = (Path(folder)/stem_name).with_suffix('.xlsx')
    ws_name = ws_name or stem_name
    wb = df_to_excel_workbook(df, wb=None, template=template, max_col_width=max_col_width,
                              ws_name=ws_name, index=index, header=header, index_filter=index_filter,
                              view_left_col=view_left_col, copy_style=(template is not None))
    wb.save(pathname)
    return pathname


def df_prefix_excel_error_prone_text(df, columns=None):
    """Append single quote to the text(s) prone to cause Excel error.
    This will append prefix (single quote) for those text cells in the DataFrame,
    so that it won't cause error when df is converted to Excel document.
    Supported prefixing(s):
    - Texts starting with '='; that will 100% cause error.
    """
    def is_errornous(value):
        if type(value) != str: return False
        if len(value) == 0: return False
        return value[0] == '=' # or value[0] == '-'

    if columns is None:
        columns = [c for c in df.columns]
    for c in columns:
        df[c] = df[c].apply(lambda x: "'"+x if is_errornous(x) else x)
